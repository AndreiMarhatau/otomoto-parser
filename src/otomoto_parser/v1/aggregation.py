#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Generate hierarchical Otomoto aggregations sheet (make → model → body_type)
from a JSONL export like results.jsonl.

Output columns (exact header):
make\tmake_count\tmake_year_range\tmake_price_range\tmake_mileage_median\tmake_registered_pct_pl
model\tmodel_count\tmodel_year_range\tmodel_price_range\tmodel_mileage_median\tmodel_registered_pct_pl
body_type\tbody_count\tbody_year_range\tbody_price_range\tbody_mileage_median\tbody_registered_pct_pl
"""

from __future__ import annotations

import argparse
import json
import os
from pathlib import Path
from typing import Any, Dict, Iterable, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.table import Table, TableStyleInfo


HEADER = [
    "make",
    "make_count",
    "make_year_range",
    "make_price_range",
    "make_mileage_median",
    "make_registered_pct_pl",
    "model",
    "model_count",
    "model_year_range",
    "model_price_range",
    "model_mileage_median",
    "model_registered_pct_pl",
    "body_type",
    "body_count",
    "body_year_range",
    "body_price_range",
    "body_mileage_median",
    "body_registered_pct_pl",
]


class AggregationError(RuntimeError):
    pass


def safe_int(value: Any) -> Optional[int]:
    try:
        if value is None:
            return None
        if isinstance(value, bool):
            return int(value)
        if isinstance(value, (int, float)):
            return int(value)
        text = str(value).strip()
        if not text:
            return None
        return int(float(text))
    except Exception:
        return None


def get_param(params: Iterable[Dict[str, Any]], key: str) -> Optional[str]:
    for param in params or []:
        if param.get("key") == key:
            return param.get("value") or param.get("displayValue")
    return None


def parse_listing(obj: Dict[str, Any]) -> Dict[str, Any]:
    node = obj.get("node") or (obj.get("edge") or {}).get("node") or {}
    params = node.get("parameters") or []

    make = get_param(params, "make")
    model = get_param(params, "model")
    body_type = get_param(params, "body_type")

    year = safe_int(get_param(params, "year"))
    mileage = safe_int(get_param(params, "mileage"))

    registered_raw = get_param(params, "registered")
    if registered_raw is None or str(registered_raw).strip() == "":
        registered = None
    else:
        normalized = str(registered_raw).strip().lower()
        registered = 1 if normalized in {"1", "true", "tak", "yes"} else 0

    price_units = None
    try:
        price_units = node.get("price", {}).get("amount", {}).get("units")
    except Exception:
        price_units = None
    price = safe_int(price_units)

    return {
        "make": make,
        "model": model,
        "body_type": body_type,
        "year": year,
        "price": price,
        "mileage": mileage,
        "registered": registered,
    }


def read_jsonl(path: Path) -> pd.DataFrame:
    rows = []
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            obj = json.loads(line)
            rows.append(parse_listing(obj))
    if not rows:
        return pd.DataFrame(columns=["make", "model", "body_type", "year", "price", "mileage", "registered"])

    df = pd.DataFrame(rows)
    df = df.dropna(subset=["make", "model", "body_type"], how="any")

    for col in ["year", "price", "mileage", "registered"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    if "registered" not in df.columns:
        df["registered"] = pd.Series(dtype="float64")

    return df


def range_str(series: pd.Series) -> str:
    values = series.dropna()
    if values.empty:
        return ""
    mn = int(values.min())
    mx = int(values.max())
    return str(mn) if mn == mx else f"{mn}-{mx}"


def median_int(series: pd.Series) -> Optional[int]:
    values = series.dropna()
    if values.empty:
        return None
    return int(round(float(values.median())))


def agg_metrics(df: pd.DataFrame) -> Dict[str, Any]:
    registered_values = df["registered"].dropna()
    registered_pct = float(registered_values.mean()) if not registered_values.empty else None
    return {
        "count": int(df.shape[0]),
        "year_range": range_str(df["year"]),
        "price_range": range_str(df["price"]),
        "mileage_median": median_int(df["mileage"]),
        "registered_pct": registered_pct,
    }


def build_hier_rows(df: pd.DataFrame) -> pd.DataFrame:
    out_rows = []

    if df.empty:
        return pd.DataFrame(columns=HEADER)

    make_counts = df.groupby("make").size().sort_values(ascending=False)
    for make, _ in make_counts.items():
        df_make = df[df["make"] == make]
        make_metrics = agg_metrics(df_make)

        model_counts = df_make.groupby("model").size().sort_values(ascending=False)
        for model, _ in model_counts.items():
            df_model = df_make[df_make["model"] == model]
            model_metrics = agg_metrics(df_model)

            body_counts = df_model.groupby("body_type").size().sort_values(ascending=False)
            for body_type, _ in body_counts.items():
                df_body = df_model[df_model["body_type"] == body_type]
                body_metrics = agg_metrics(df_body)

                out_rows.append(
                    {
                        "make": make,
                        "make_count": make_metrics["count"],
                        "make_year_range": make_metrics["year_range"],
                        "make_price_range": make_metrics["price_range"],
                        "make_mileage_median": make_metrics["mileage_median"],
                        "make_registered_pct_pl": make_metrics["registered_pct"],
                        "model": model,
                        "model_count": model_metrics["count"],
                        "model_year_range": model_metrics["year_range"],
                        "model_price_range": model_metrics["price_range"],
                        "model_mileage_median": model_metrics["mileage_median"],
                        "model_registered_pct_pl": model_metrics["registered_pct"],
                        "body_type": body_type,
                        "body_count": body_metrics["count"],
                        "body_year_range": body_metrics["year_range"],
                        "body_price_range": body_metrics["price_range"],
                        "body_mileage_median": body_metrics["mileage_median"],
                        "body_registered_pct_pl": body_metrics["registered_pct"],
                    }
                )

    return pd.DataFrame(out_rows, columns=HEADER)


def autosize_columns(ws, min_width: int = 10, max_width: int = 45) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        width = max(min_width, min(max_width, max_len + 2))
        ws.column_dimensions[col_letter].width = width


def write_excel(df_out: pd.DataFrame, out_path: Path, sheet_name: str = "Aggregations") -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    if out_path.exists():
        out_path.unlink()

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df_out.to_excel(writer, index=False, sheet_name=sheet_name)

    wb = load_workbook(out_path)
    ws = wb[sheet_name]

    ws.freeze_panes = "A2"

    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

    ws.auto_filter.ref = ws.dimensions

    col_index = {name: i + 1 for i, name in enumerate(HEADER)}
    pct_cols = ["make_registered_pct_pl", "model_registered_pct_pl", "body_registered_pct_pl"]
    int_cols = [
        "make_count",
        "model_count",
        "body_count",
        "make_mileage_median",
        "model_mileage_median",
        "body_mileage_median",
    ]

    for row in range(2, ws.max_row + 1):
        for col_name in int_cols:
            cell = ws.cell(row=row, column=col_index[col_name])
            if cell.value not in (None, ""):
                cell.number_format = "0"
        for col_name in pct_cols:
            cell = ws.cell(row=row, column=col_index[col_name])
            if cell.value not in (None, ""):
                cell.number_format = "0%"

    if ws.max_row >= 2 and ws.max_column >= 1:
        table = Table(displayName="AggregationsTable", ref=ws.dimensions)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)

    autosize_columns(ws)

    wb.save(out_path)


def default_output_path(input_path: Path) -> Path:
    return input_path.with_name(f"{input_path.stem}_aggregations.xlsx")


def generate_aggregations(input_path: Path, output_path: Optional[Path] = None) -> Path:
    if not input_path.exists():
        raise AggregationError(f"Input file does not exist: {input_path}")

    output_path = output_path or default_output_path(input_path)
    if output_path.parent.exists():
        if not os.access(output_path.parent, os.W_OK):
            if output_path == default_output_path(input_path):
                output_path = Path.cwd() / output_path.name
            else:
                raise AggregationError(f"Output directory is not writable: {output_path.parent}")
    else:
        try:
            output_path.parent.mkdir(parents=True, exist_ok=True)
        except OSError as exc:
            if output_path == default_output_path(input_path):
                output_path = Path.cwd() / output_path.name
            else:
                raise AggregationError(
                    f"Unable to create output directory: {output_path.parent}"
                ) from exc

    df = read_jsonl(input_path)
    df_out = build_hier_rows(df)
    write_excel(df_out, output_path)
    return output_path


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Generate aggregation sheet from results JSONL")
    parser.add_argument("--input", "-i", required=True, help="Path to results.jsonl")
    parser.add_argument(
        "--output",
        "-o",
        default=None,
        help="Output xlsx path (defaults to <results>_aggregations.xlsx)",
    )
    return parser


def main() -> None:
    parser = build_arg_parser()
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output) if args.output else None

    final_path = generate_aggregations(input_path, output_path)
    print(str(final_path))


__all__ = [
    "AggregationError",
    "build_arg_parser",
    "build_hier_rows",
    "default_output_path",
    "generate_aggregations",
    "main",
    "parse_listing",
    "read_jsonl",
    "write_excel",
]
