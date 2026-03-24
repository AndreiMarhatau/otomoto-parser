from __future__ import annotations

import os
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.table import Table, TableStyleInfo

from ._aggregation_common import AggregationError, HEADER
from ._aggregation_metrics import build_hier_rows
from ._aggregation_records import read_jsonl


def autosize_columns(worksheet, min_width: int = 10, max_width: int = 45) -> None:
    for column in worksheet.columns:
        width = max(len("" if cell.value is None else str(cell.value)) for cell in column)
        worksheet.column_dimensions[column[0].column_letter].width = max(min_width, min(max_width, width + 2))


def write_excel(frame: pd.DataFrame, out_path: Path, sheet_name: str = "Aggregations") -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    if out_path.exists():
        out_path.unlink()
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        frame.to_excel(writer, index=False, sheet_name=sheet_name)
    workbook = load_workbook(out_path)
    worksheet = workbook[sheet_name]
    _format_worksheet(worksheet)
    workbook.save(out_path)


def _format_worksheet(worksheet) -> None:
    worksheet.freeze_panes = "A2"
    _format_headers(worksheet)
    _format_body_cells(worksheet)
    if worksheet.max_row >= 2 and worksheet.max_column >= 1:
        table = Table(displayName="AggregationsTable", ref=worksheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)
    worksheet.auto_filter.ref = worksheet.dimensions
    autosize_columns(worksheet)


def _format_headers(worksheet) -> None:
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)


def _format_body_cells(worksheet) -> None:
    column_index = {name: index + 1 for index, name in enumerate(HEADER)}
    int_columns = [
        "make_count",
        "model_count",
        "body_count",
        "make_mileage_median",
        "model_mileage_median",
        "body_mileage_median",
    ]
    pct_columns = ["make_registered_pct_pl", "model_registered_pct_pl", "body_registered_pct_pl"]
    for row in range(2, worksheet.max_row + 1):
        _apply_number_format(worksheet, row, column_index, (int_columns, "0"))
        _apply_number_format(worksheet, row, column_index, (pct_columns, "0%"))


def _apply_number_format(worksheet, row: int, column_index: dict[str, int], format_spec: tuple[list[str], str]) -> None:
    columns, pattern = format_spec
    for column_name in columns:
        cell = worksheet.cell(row=row, column=column_index[column_name])
        if cell.value not in (None, ""):
            cell.number_format = pattern


def default_output_path(input_path: Path) -> Path:
    return input_path.with_name(f"{input_path.stem}_aggregations.xlsx")


def generate_aggregations(input_path: Path, output_path: Path | None = None) -> Path:
    if not input_path.exists():
        raise AggregationError(f"Input file does not exist: {input_path}")

    resolved_output = output_path or default_output_path(input_path)
    resolved_output = _ensure_output_path(input_path, resolved_output)
    write_excel(build_hier_rows(read_jsonl(input_path)), resolved_output)
    return resolved_output


def _ensure_output_path(input_path: Path, output_path: Path) -> Path:
    if output_path.parent.exists():
        if os.access(output_path.parent, os.W_OK):
            return output_path
        if output_path == default_output_path(input_path):
            return Path.cwd() / output_path.name
        raise AggregationError(f"Output directory is not writable: {output_path.parent}")
    try:
        output_path.parent.mkdir(parents=True, exist_ok=True)
    except OSError as exc:
        if output_path == default_output_path(input_path):
            return Path.cwd() / output_path.name
        raise AggregationError(f"Unable to create output directory: {output_path.parent}") from exc
    return output_path
