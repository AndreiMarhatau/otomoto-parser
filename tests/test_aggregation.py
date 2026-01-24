from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from otomoto_parser.aggregation import (
    build_hier_rows,
    default_output_path,
    generate_aggregations,
    read_jsonl,
)


FIXTURE_PATH = Path(__file__).parent / "fixtures" / "aggregation_sample.jsonl"


def test_read_jsonl_parses_fixture() -> None:
    df = read_jsonl(FIXTURE_PATH)

    assert df.shape[0] == 2
    assert set(df["make"]) == {"audi", "toyota"}
    assert set(df["model"]) == {"q5", "avensis"}
    assert set(df["body_type"]) == {"suv", "combi"}


def test_build_hier_rows_single_row_per_make() -> None:
    df = read_jsonl(FIXTURE_PATH)
    out = build_hier_rows(df)

    assert list(out.columns) == [
        "make",
        "make_count",
        "make_year_range",
        "make_price_range",
        "make_mileage_median",
        "make_registered_pct_pl",
        "model",
        "model_count",
        "model_year_range",
        "model_price_range",
        "model_mileage_median",
        "model_registered_pct_pl",
        "body_type",
        "body_count",
        "body_year_range",
        "body_price_range",
        "body_mileage_median",
        "body_registered_pct_pl",
    ]
    assert out.shape[0] == 2

    by_make = {row.make: row for row in out.itertuples(index=False)}
    audi = by_make["audi"]
    toyota = by_make["toyota"]

    assert audi.make_count == 1
    assert audi.make_year_range == "2012"
    assert audi.make_price_range == "41000"
    assert audi.make_mileage_median == 253000
    assert audi.make_registered_pct_pl == pytest.approx(1.0)
    assert audi.model == "q5"
    assert audi.body_type == "suv"

    assert toyota.make_count == 1
    assert toyota.make_year_range == "2014"
    assert toyota.make_price_range == "43000"
    assert toyota.make_mileage_median == 150917
    assert toyota.make_registered_pct_pl == pytest.approx(1.0)
    assert toyota.model == "avensis"
    assert toyota.body_type == "combi"


def test_generate_aggregations_creates_excel(tmp_path: Path) -> None:
    input_path = tmp_path / "results.jsonl"
    input_path.write_text(FIXTURE_PATH.read_text(encoding="utf-8"), encoding="utf-8")

    output_path = generate_aggregations(input_path)
    assert output_path == default_output_path(input_path)
    assert output_path.exists()

    wb = load_workbook(output_path)
    assert "Aggregations" in wb.sheetnames
    ws = wb["Aggregations"]
    header = [cell.value for cell in ws[1]]
    assert header[0] == "make"
    assert header[-1] == "body_registered_pct_pl"

    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(data_rows) == 2


def test_generate_aggregations_overwrites_existing(tmp_path: Path) -> None:
    input_path = tmp_path / "results.jsonl"
    input_path.write_text(FIXTURE_PATH.read_text(encoding="utf-8"), encoding="utf-8")
    output_path = default_output_path(input_path)
    output_path.write_text("old", encoding="utf-8")

    generate_aggregations(input_path, output_path)
    assert output_path.exists()
    assert output_path.read_bytes() != b"old"


def test_build_hier_rows_handles_empty() -> None:
    df = pd.DataFrame(columns=["make", "model", "body_type", "year", "price", "mileage", "registered"])
    out = build_hier_rows(df)
    assert out.empty
    assert list(out.columns) == [
        "make",
        "make_count",
        "make_year_range",
        "make_price_range",
        "make_mileage_median",
        "make_registered_pct_pl",
        "model",
        "model_count",
        "model_year_range",
        "model_price_range",
        "model_mileage_median",
        "model_registered_pct_pl",
        "body_type",
        "body_count",
        "body_year_range",
        "body_price_range",
        "body_mileage_median",
        "body_registered_pct_pl",
    ]
